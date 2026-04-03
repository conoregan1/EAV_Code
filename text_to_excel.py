"""
ESP32 Sensor Logger — Binary Data Processor v1.1.1
====================================================
Reads data.bin (binary format) from the SD card and appends records
to a dated .xlsx file. Time column is written as plain text
(HH:MM:SS.mmm) so Excel never auto-converts it to a time serial and
drops the milliseconds.

If data.bin is not found (e.g. SD card not plugged in), the script
skips the data import step and returns cleanly so downstream analysis
scripts can still load and process the existing .xlsx file.

Binary record format (38 bytes, little-endian, packed):
  [0]      uint8   sync byte (0xAA)
  [1–4]    uint32  time_ms  — milliseconds since midnight (0–86399999)
  [5]      uint8   gps_valid  (1 = valid fix, 0 = no fix)
  [6–9]    int32   latitude  × 1,000,000
  [10–13]  int32   longitude × 1,000,000
  [14–17]  float32 accelX
  [18–21]  float32 accelY
  [22–25]  float32 accelZ
  [26–29]  float32 roll
  [30–33]  float32 pitch
  [34–37]  float32 yaw

File format:
  Bytes 0–3 : magic "ESPB"
  Bytes 4+  : records, back to back

Output columns:
  Time, Lat, Long, x-axis, y-axis, z-axis, Roll, Pitch, Yaw
"""

import struct
import os
import sys
import time
from datetime import datetime

import openpyxl

# ===================================================================
# Binary format constants — must match Arduino sketch exactly
# ===================================================================
FILE_MAGIC    = b'ESPB'
SYNC_BYTE     = 0xAA

RECORD_FORMAT = '<BIBiiffffff'
RECORD_SIZE   = struct.calcsize(RECORD_FORMAT)

assert RECORD_SIZE == 38, (
    f"RECORD_SIZE should be 38 but got {RECORD_SIZE}. "
    "Check the struct format string."
)


def ms_to_timestamp(ms):
    """Convert milliseconds since midnight to HH:MM:SS.mmm string."""
    ms = int(ms)
    h  = ms // 3_600_000;  ms %= 3_600_000
    m  = ms // 60_000;     ms %= 60_000
    s  = ms // 1_000;      ms %= 1_000
    return f'{h:02d}:{m:02d}:{s:02d}.{ms:03d}'


def print_progress(current, total, label='Writing', bar_width=40):
    """Print a single-line updating progress bar to the terminal."""
    pct     = current / total if total > 0 else 0
    filled  = int(bar_width * pct)
    bar     = '=' * filled + '-' * (bar_width - filled)
    elapsed = time.time() - print_progress.start_time
    line    = f'\r{label:<10} [{bar}] {current}/{total} ({pct*100:5.1f}%) {elapsed:.1f}s'
    sys.stdout.write(line)
    sys.stdout.flush()
    if current >= total:
        sys.stdout.write('\n')
        sys.stdout.flush()

print_progress.start_time = 0


def process_data_file(data_file_path, output_directory):
    """
    Reads data.bin, parses binary records, appends to today's .xlsx
    file with Time column stored as text, then wipes the source file.

    If data.bin does not exist (SD card not plugged in), prints a
    clear message and returns without error so the calling script can
    continue with analysis of any existing .xlsx data.
    """

    # ------------------------------------------------------------------
    # 0. Check whether the SD card / data file is present at all.
    #    Return gracefully if not — don't crash the analysis script.
    # ------------------------------------------------------------------
    if not os.path.exists(data_file_path):
        print(f"Note: Data file not found at '{data_file_path}'.")
        print("      SD card may not be plugged in — skipping import.")
        print("      Continuing with existing .xlsx data if available.")
        return

    # ------------------------------------------------------------------
    # 1. Read the whole binary file
    # ------------------------------------------------------------------
    try:
        with open(data_file_path, 'rb') as f:
            raw = f.read()
    except PermissionError:
        print(f"Error: Cannot read '{data_file_path}' — check SD card permissions.")
        return
    except OSError as e:
        print(f"Error reading data file: {e}")
        return

    if len(raw) < 4:
        print("Data file is empty or too small. Nothing to import.")
        return

    # ------------------------------------------------------------------
    # 2. Validate and skip 4-byte file header
    # ------------------------------------------------------------------
    if raw[:4] == FILE_MAGIC:
        offset = 4
        print(f"File header OK. Total bytes: {len(raw)}, "
              f"Records expected: {(len(raw) - 4) // RECORD_SIZE}")
    else:
        print("Warning: File magic 'ESPB' not found. Reading from byte 0.")
        offset = 0

    if len(raw) - offset < RECORD_SIZE:
        print("No complete records found in data file.")
        return

    # ------------------------------------------------------------------
    # 3. Parse binary records
    # ------------------------------------------------------------------
    total_expected = (len(raw) - offset) // RECORD_SIZE
    print(f"Parsing {total_expected} records...")

    rows           = []
    records_parsed = 0
    bytes_skipped  = 0

    print_progress.start_time = time.time()

    while offset + RECORD_SIZE <= len(raw):
        if raw[offset] != SYNC_BYTE:
            offset        += 1
            bytes_skipped += 1
            continue

        record_bytes = raw[offset : offset + RECORD_SIZE]
        (sync, time_ms, gps_valid,
         lat_e6, lon_e6,
         accel_x, accel_y, accel_z,
         roll, pitch, yaw) = struct.unpack(RECORD_FORMAT, record_bytes)

        time_str = ms_to_timestamp(time_ms)
        lat = round(lat_e6 / 1_000_000, 6) if gps_valid else None
        lon = round(lon_e6 / 1_000_000, 6) if gps_valid else None

        rows.append([
            time_str,
            lat,
            lon,
            round(float(accel_x), 4),
            round(float(accel_y), 4),
            round(float(accel_z), 4),
            round(float(roll),   2),
            round(float(pitch),  2),
            round(float(yaw),    2),
        ])

        records_parsed += 1
        offset         += RECORD_SIZE

        if records_parsed % 250 == 0 or records_parsed == total_expected:
            print_progress(records_parsed, total_expected, label='Parsing')

    print_progress(records_parsed, total_expected, label='Parsing')

    if bytes_skipped > 0:
        print(f"Warning: Skipped {bytes_skipped} byte(s) while resyncing.")

    if not rows:
        print("No valid records found. Nothing to import.")
        return

    print(f"Parsed {records_parsed} records successfully.")

    # ------------------------------------------------------------------
    # 4. Write / append to dated .xlsx
    # ------------------------------------------------------------------
    os.makedirs(output_directory, exist_ok=True)

    current_date = datetime.now().strftime('%Y-%m-%d')
    xlsx_path    = os.path.join(output_directory, f'{current_date}.xlsx')

    HEADERS = ['Time', 'Lat', 'Long', 'x-axis', 'y-axis', 'z-axis',
               'Roll', 'Pitch', 'Yaw']

    if os.path.exists(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        existing_rows = ws.max_row - 1
        print(f"Appending {len(rows)} records to existing file "
              f"({existing_rows} rows already present): {xlsx_path}")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = current_date
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        print(f"Creating new file: {xlsx_path}")

    total_rows = len(rows)
    print_progress.start_time = time.time()

    for idx, row_data in enumerate(rows, start=1):
        ws.append(row_data)
        time_cell               = ws.cell(row=ws.max_row, column=1)
        time_cell.value         = row_data[0]
        time_cell.number_format = '@'
        time_cell.data_type     = 's'
        if idx % 250 == 0 or idx == total_rows:
            print_progress(idx, total_rows, label='Writing')

    ws.column_dimensions['A'].width = 16
    wb.save(xlsx_path)
    print(f"Saved. Total rows in file: {ws.max_row - 1}")

    # ------------------------------------------------------------------
    # 5. Wipe the source binary file
    # ------------------------------------------------------------------
    with open(data_file_path, 'wb') as f:
        f.write(b'')
    print(f"Wiped source file: {data_file_path}")


# ===================================================================
# Entry point
# ===================================================================
if __name__ == "__main__":
    data_file_path   = 'D:/data.bin'
    output_directory = 'C:/Users/conor/OneDrive/Documents/EAV_Data'
    process_data_file(data_file_path, output_directory)
